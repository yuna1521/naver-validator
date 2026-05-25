import os
import re
import requests
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 설정
# ============================================================
load_dotenv()
CLIENT_ID = os.getenv("NAVER_CLIENT_ID")
CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET")

INPUT_FILE = "테스트_정관장.xlsx"
OUTPUT_FILE = f"검수결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

TEST_MODE = True
TEST_LIMIT = 10

if not CLIENT_ID or not CLIENT_SECRET:
    print("ERROR: API 키 누락. .env 파일을 확인하세요.")
    exit()


# ============================================================
# 헬퍼 함수들
# ============================================================
def clean_text(text):
    return re.sub(r"<[^>]+>", "", str(text)).strip()


def extract_keywords(product_name):
    return [t for t in product_name.split() if len(t) >= 2]


def extract_units(product_name):
    pattern = r"(\d+(?:\.\d+)?)\s*(ml|l|g|kg|포|개|정|캡슐|환|매|입|병|봉|팩)"
    matches = re.findall(pattern, product_name, flags=re.IGNORECASE)
    return [f"{num}{unit.lower()}" for num, unit in matches]


def match_score(product_name, item_title):
    title_clean = clean_text(item_title)
    원본_단위들 = extract_units(product_name)
    제목_단위들 = extract_units(title_clean)
    for unit in 원본_단위들:
        if unit not in 제목_단위들:
            return 0.0
    keywords = extract_keywords(product_name)
    if not keywords:
        return 0.0
    matched = sum(1 for kw in keywords if kw in title_clean)
    return matched / len(keywords)


# ============================================================
# 검수 함수 1: 산식 검증
# ============================================================
def validate_formula(원가, 수수료, 판매가):
    if pd.isna(원가) or pd.isna(수수료) or pd.isna(판매가):
        return "데이터 누락", "원가/수수료/판매가 중 빈 값 있음"
    
    expected_수수료 = 원가 * 0.15
    expected_판매가 = 원가 + 수수료
    
    수수료_정확함 = abs(수수료 - expected_수수료) <= 1
    판매가_정확함 = abs(판매가 - expected_판매가) <= 1
    
    if 수수료_정확함 and 판매가_정확함:
        return "PASS", ""
    
    메모_조각 = []
    if not 수수료_정확함:
        실제_수수료율 = (수수료 / 원가) * 100
        메모_조각.append(f"수수료 {실제_수수료율:.1f}%")
    if not 판매가_정확함:
        차액 = 판매가 - expected_판매가
        방향 = "높음" if 차액 > 0 else "낮음"
        메모_조각.append(f"판매가가 원가+수수료보다 {방향} ({int(abs(차액)):,}원)")
    
    결과 = "수수료율 불일치" if not 수수료_정확함 else "판매가 불일치"
    return 결과, " / ".join(메모_조각)


# ============================================================
# 검수 함수 2: 네이버 가격 조회
# ============================================================
def search_naver_price(product_name, brand_name):
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {"X-Naver-Client-Id": CLIENT_ID, "X-Naver-Client-Secret": CLIENT_SECRET}
    params = {"query": product_name, "display": 10, "sort": "asc"}
    
    try:
        response = requests.get(url, headers=headers, params=params, timeout=10)
    except requests.exceptions.RequestException as e:
        return {"status": "ERROR", "price": None, "message": f"네트워크 오류: {e}"}
    
    if response.status_code == 429:
        return {"status": "API_LIMIT", "price": None, "message": "API 일일 한도 도달"}
    if response.status_code != 200:
        return {"status": "ERROR", "price": None, "message": f"API 오류 {response.status_code}"}
    
    items = response.json().get("items", [])
    if not items:
        return {"status": "NO_MATCH", "price": None, "message": "네이버 검색 결과 0건"}
    
    candidates = []
    for item in items:
        if item.get("productType") == "1":
            continue
        item_brand = clean_text(item.get("brand", ""))
        if item_brand and brand_name and item_brand != brand_name:
            continue
        candidates.append(item)
    
    if not candidates:
        return {"status": "NO_MATCH", "price": None, "message": "조건 맞는 상품 없음"}
    
    scored = [{"item": i, "score": match_score(product_name, i["title"]), "price": int(i["lprice"])} for i in candidates]
    matched = [s for s in scored if s["score"] >= 0.7]
    
    if not matched:
        best = max(scored, key=lambda s: s["score"])
        return {"status": "NO_MATCH", "price": None, "message": f"매칭 신뢰도 낮음 ({best['score']:.0%})"}
    
    best_match = min(matched, key=lambda s: s["price"])
    return {"status": "OK", "price": best_match["price"], "message": "정상"}


# ============================================================
# 메인 처리
# ============================================================
def main():
    print(f"엑셀 파일 읽기: {INPUT_FILE}")
    df = pd.read_excel(INPUT_FILE)
    df = df.dropna(subset=["상품명"]).reset_index(drop=True)
    
    print(f"검수 대상: {len(df)}개 상품")
    
    if TEST_MODE and len(df) > TEST_LIMIT:
        print(f"⚠️ 테스트 모드: 처음 {TEST_LIMIT}건만 처리합니다.")
        df = df.head(TEST_LIMIT)
    
    answer = input(f"검수를 시작할까요? (y/n): ")
    if answer.lower() != "y":
        print("취소되었습니다.")
        return
    
    print("\n검수 시작...")
    results = []
    
    for index, row in df.iterrows():
        print(f"  [{index+1}/{len(df)}] {row['상품명']}")
        
        원가 = row["원가"]
        수수료 = row["플랫폼 수수료 (15%)"]
        판매가 = row["판매가"]
        브랜드명 = str(row.get("브랜드명", "")).strip()
        
        # 검수 1: 산식
        산식결과, 산식메모 = validate_formula(원가, 수수료, 판매가)
        
        # 검수 2: 네이버 가격
        naver_result = search_naver_price(str(row["상품명"]), 브랜드명)
        
        if naver_result["status"] == "API_LIMIT":
            print("\n⚠️ API 한도 도달. 검수를 중단하고 지금까지의 결과를 저장합니다.")
            break
        
        네이버조회가 = naver_result["price"]
        
        # 차이 계산
        기재_네이버최저가 = row.get("네이버최저가")
        if 네이버조회가 and not pd.isna(기재_네이버최저가):
            차액 = int(기재_네이버최저가) - 네이버조회가
            차이율 = (차액 / 네이버조회가) * 100 if 네이버조회가 else 0
            차이표시 = f"{차액:+,} ({차이율:+.1f}%)"
        else:
            차이표시 = "—"
        
        # 종합 메모 만들기
        메모조각 = []
        if 산식메모:
            메모조각.append(산식메모)
        if naver_result["status"] != "OK":
            메모조각.append(naver_result["message"])
        if naver_result["status"] == "OK" and 네이버조회가 and not pd.isna(판매가):
            if 판매가 > 네이버조회가:
                차이율 = ((판매가 - 네이버조회가) / 네이버조회가) * 100
                역산원가 = int(네이버조회가 / 1.15)
                메모조각.append(
                    f"판매가 {int(판매가):,}원이 네이버 조회가 {네이버조회가:,}원보다 {차이율:.1f}% 비쌈. 원가 {역산원가:,}원 이하로 협상 필요."
                )
        
        검수메모 = " / ".join(메모조각) if 메모조각 else ""
        
        # 종합 결과
        if (산식결과 == "PASS" and naver_result["status"] == "OK"
                and (pd.isna(판매가) or not 네이버조회가 or 판매가 <= 네이버조회가)):
            종합결과 = "승인 가능"
        else:
            종합결과 = "확인 필요"
        
        results.append({
            "종합결과": 종합결과,
            "산식검증": 산식결과,
            "네이버조회가": 네이버조회가 if 네이버조회가 else "—",
            "차이": 차이표시,
            "검수메모": 검수메모 if 검수메모 else "—",
        })
    
    df = df.head(len(results))
    
    for col in ["종합결과", "산식검증", "네이버조회가", "차이", "검수메모"]:
        df[col] = [r[col] for r in results]
    
    save_excel(df, OUTPUT_FILE)
    
    승인_가능 = sum(1 for r in results if r["종합결과"] == "승인 가능")
    확인_필요 = len(results) - 승인_가능
    print(f"\n=== 검수 완료 ===")
    print(f"  전체: {len(results)}건")
    print(f"  승인 가능: {승인_가능}건")
    print(f"  확인 필요: {확인_필요}건")
    print(f"  결과 파일: {OUTPUT_FILE}")


# ============================================================
# 엑셀 저장
# ============================================================
def save_excel(df, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "검수결과"
    
    summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    check_header_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    ok_fill = PatternFill(start_color="EAF3DE", end_color="EAF3DE", fill_type="solid")
    review_fill = PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    승인_가능 = (df["종합결과"] == "승인 가능").sum()
    확인_필요 = (df["종합결과"] == "확인 필요").sum()
    검수시각 = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    ws.cell(row=1, column=1, value=f"[검수 요약]  전체 {len(df)}건  |  승인 가능 {승인_가능}건  |  확인 필요 {확인_필요}건  |  검수시각 {검수시각}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    summary_cell = ws.cell(row=1, column=1)
    summary_cell.fill = summary_fill
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 25
    
    원본_컬럼수 = 8
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = check_header_fill if col_idx > 원본_컬럼수 else header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        is_review = row["종합결과"] == "확인 필요"
        row_fill = review_fill if is_review else ok_fill
        
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 원본_컬럼수:
                cell.fill = row_fill
            if df.columns[col_idx-1] == "종합결과":
                cell.font = Font(bold=True)
    
    widths = {
        "No": 5, "브랜드명": 10, "상품명": 35, "원가": 12, 
        "플랫폼 수수료 (15%)": 16, "판매가": 12, "네이버최저가": 12, "네이버최저가 URL": 30,
        "종합결과": 11, "산식검증": 14, "네이버조회가": 13, "차이": 18, "검수메모": 50,
    }
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)
    
    ws.freeze_panes = "D4"
    
    wb.save(filename)


if __name__ == "__main__":
    main()