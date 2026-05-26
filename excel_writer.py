"""
원본 엑셀 파일을 그대로 보존하면서 검수 결과를 우측 컬럼에 추가하여 저장.
새 양식(애드웰 상품제안서) 전용.
"""
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 검수 결과 컬럼들 (Q열부터 추가됨)
RESULT_COLUMNS = ["결과", "네이버 조회가", "차이 (A − B)", "시장 풍부도", "검수 메모"]

# 새 양식 원본의 컬럼 수 (A~P, 총 16개)
ORIG_COL_COUNT = 16

# 새 양식 데이터 시작 행 (1행 제목, 2행 헤더, 3행부터 데이터)
DATA_START_ROW = 3


def append_results_to_excel(original_file_bytes, results):
    """원본 엑셀 파일의 첫 시트에 검수 결과 컬럼을 추가하여 bytes로 반환.
    
    원본의 시트, 서식, 수식, 셀 병합은 모두 보존된다.
    검수 결과는 Q열(17번째)부터 시작.
    
    인자:
        original_file_bytes: 원본 엑셀의 BytesIO 또는 bytes
        results: validate_row가 반환한 dict 리스트
    
    반환: 변경된 엑셀의 bytes
    """
    # 원본 로드
    if isinstance(original_file_bytes, bytes):
        original_file_bytes = BytesIO(original_file_bytes)
    
    wb = load_workbook(original_file_bytes)
    ws = wb.worksheets[0]  # 첫 시트
    
    # 스타일 정의
    header_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    ok_fill = PatternFill(start_color="EAF3DE", end_color="EAF3DE", fill_type="solid")
    review_fill = PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    start_col = ORIG_COL_COUNT + 1  # Q열 = 17
    
    # 1) 헤더 행 (2행)
    for i, header in enumerate(RESULT_COLUMNS):
        cell = ws.cell(row=2, column=start_col + i, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 2) 데이터 행 (3행부터)
    for i, result in enumerate(results):
        row_idx = DATA_START_ROW + i
        fill = ok_fill if result["결과"] == "승인 가능" else review_fill
        
        values = [
            result["결과"],
            result["네이버 조회가"],
            result["차이"],
            result["시장 풍부도"],
            result["검수 메모"],
        ]
        
        for j, val in enumerate(values):
            cell = ws.cell(row=row_idx, column=start_col + j, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if j == 0:  # 결과 컬럼 굵게
                cell.font = Font(bold=True)
    
    # 3) 컬럼 너비
    widths = [12, 14, 18, 22, 50]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w
    
    # 4) 행 높이 (검수 결과 행)
    for i in range(len(results)):
        row_idx = DATA_START_ROW + i
        ws.row_dimensions[row_idx].height = 60
    
    # 5) 틀 고정: 2행(헤더)과 E열(상품명)까지 고정
    ws.freeze_panes = "F3"
    
    # bytes로 변환
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()